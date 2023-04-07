import tkinter as tk
import openpyxl
from xlrd import open_workbook

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Armazenar número em Excel")
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        self.label = tk.Label(self, text="Insira um número:")
        self.label.pack()

        self.entry = tk.Entry(self)
        self.entry.pack()

        self.button = tk.Button(self, text="OK", command=self.store_number)
        self.button.pack()

    def store_number(self):
        number = self.entry.get()
        wb = openpyxl.load_workbook("banco.xlsx")
        ws = wb.active
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=number)
        wb.save("banco.xlsx")
        self.entry.delete(0, tk.END)

root = tk.Tk()
app = Application(master=root)
app.mainloop()
